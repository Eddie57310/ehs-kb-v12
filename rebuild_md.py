"""
rebuild_md.py  v1.0
从 Local_KB 源文件生成干净的 Markdown，存入 reviewed_md/

处理范围：国家规定 / 公司内部
  PDF  → MinerU → content_list.json → MD（表格/公式保留为文本）
  DOCX → LibreOffice HTML → BeautifulSoup → MD（序号/页眉/表格完整保留）
  XLSX → openpyxl → MD
跳过：EHS案例（走 Vision LLM 原方案）
"""

import os, json, re, glob, subprocess, logging, time
from pathlib import Path
from html.parser import HTMLParser

BASE_DIR    = os.path.expanduser("~/doc_parser_v12")
KB_DIR      = f"{BASE_DIR}/Local_KB"
MD_DIR      = f"{BASE_DIR}/reviewed_md"
CACHE_DIR   = f"{BASE_DIR}/batch_output"
LOG_DIR     = f"{BASE_DIR}/logs"
MAGIC_PDF   = f"{BASE_DIR}/venv/bin/magic-pdf"

SKIP_DOMAINS = {"EHS案例"}
# 部门域下"案例"子目录也走 Vision LLM，跳过 MD 解析
SKIP_SUBDIRS = {"案例"}

os.makedirs(MD_DIR,    exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)
os.makedirs(LOG_DIR,   exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(
            f"{LOG_DIR}/rebuild_md_{time.strftime('%Y%m%d_%H%M%S')}.log",
            encoding="utf-8",
        ),
    ],
)
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════
#  1. HTML 表格 → Markdown
# ═══════════════════════════════════════════════════════

class _TblParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []

    def handle_endtag(self, tag):
        if tag == "tr" and self._row is not None:
            if any(c.strip() for c in self._row):
                self.rows.append(self._row)
            self._row = None
        elif tag in ("td", "th") and self._row is not None and self._cell is not None:
            self._row.append(" ".join(self._cell).strip())
            self._cell = None

    def handle_data(self, data):
        if self._cell is not None:
            t = data.strip()
            if t:
                self._cell.append(t)


def _html_table_to_md(html: str, caption: str = "") -> str:
    p = _TblParser()
    p.feed(html)
    rows = p.rows
    if not rows:
        return ""

    ncols = max(len(r) for r in rows)
    rows = [r + [""] * (ncols - len(r)) for r in rows]

    def _esc(s: str) -> str:
        return s.replace("|", "｜").replace("\n", " ")

    lines = []
    if caption:
        lines.append(f"**{caption}**")
    lines.append("| " + " | ".join(_esc(c) for c in rows[0]) + " |")
    lines.append("|" + "|".join("---" for _ in rows[0]) + "|")
    for row in rows[1:]:
        lines.append("| " + " | ".join(_esc(c) for c in row) + " |")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════
#  2. PDF → MinerU → content_list.json → MD
# ═══════════════════════════════════════════════════════

def _find_content_list(cache_dir: str, base: str) -> str | None:
    """在 cache_dir/base/ 下查找 content_list.json（兼容 MinerU 各版本目录结构）。"""
    # 标准路径：cache_dir/base/auto/base_content_list.json
    standard = os.path.join(cache_dir, base, "auto", f"{base}_content_list.json")
    if os.path.exists(standard):
        return standard
    # 模糊搜索：仅在该文件的专属子目录内查找，避免跨文件污染
    base_subdir = os.path.join(cache_dir, base)
    pattern = os.path.join(base_subdir, "**", f"*_content_list.json")
    found = glob.glob(pattern, recursive=True)
    if found:
        return found[0]
    return None


def _run_mineru(pdf_path: str, cache_dir: str) -> str | None:
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    cl = _find_content_list(cache_dir, base)
    if cl:
        logger.info(f"    cache hit: {base}")
        return cl

    logger.info(f"    MinerU 解析: {os.path.basename(pdf_path)}")
    try:
        r = subprocess.run(
            [MAGIC_PDF, "-p", pdf_path, "-o", cache_dir, "-m", "auto"],
            capture_output=True, text=True, timeout=3600,
        )
        cl = _find_content_list(cache_dir, base)
        if cl:
            return cl
        logger.error(f"    MinerU 未生成 content_list: {r.stderr[-200:]}")
    except subprocess.TimeoutExpired:
        logger.error(f"    MinerU 超时")
    except Exception as e:
        logger.error(f"    MinerU 异常: {e}")
    return None


def _content_list_to_md(cl_path: str) -> str:
    with open(cl_path, encoding="utf-8") as f:
        items = json.load(f)

    blocks = []

    for item in items:
        t = item.get("type", "text")

        if t == "text":
            text = item.get("text", "").strip()
            if not text:
                continue
            level = item.get("text_level")
            # text_level 可能是 int（1-4）或字符串（"h1"-"h4"）
            if isinstance(level, str) and level.startswith("h"):
                try:
                    level = int(level[1:])
                except ValueError:
                    level = None
            if isinstance(level, int) and 1 <= level <= 4 and len(text) >= 3:
                blocks.append("#" * level + " " + text)
            else:
                blocks.append(text)

        elif t == "table":
            captions = item.get("table_caption", [])
            caption  = " ".join(captions).strip() if captions else ""
            body     = item.get("table_body", "")
            if body:
                md = _html_table_to_md(body, caption)
                if md:
                    blocks.append(md)
                    continue
            # 无 body 则只保留标题
            if caption:
                blocks.append(f"**{caption}**（表格内容缺失）")

        elif t in ("equation", "interline_equation"):
            eq = (item.get("latex") or item.get("text") or "").strip()
            if eq:
                blocks.append(f"$$\n{eq}\n$$")

        elif t in ("image", "figure"):
            captions = item.get("img_caption", [])
            caption  = " ".join(captions).strip() if captions else ""
            if caption:
                blocks.append(f"<!-- 图片：{caption} -->")

    return "\n\n".join(b for b in blocks if b.strip())


def pdf_to_md(pdf_path: str, rel_dir: str) -> str | None:
    cache_subdir = os.path.join(CACHE_DIR, rel_dir)
    os.makedirs(cache_subdir, exist_ok=True)
    cl = _run_mineru(pdf_path, cache_subdir)
    if not cl:
        return None
    try:
        return _content_list_to_md(cl)
    except Exception as e:
        logger.error(f"    content_list 解析失败: {e}")
        return None


# ═══════════════════════════════════════════════════════
#  3. DOCX → MD
# ═══════════════════════════════════════════════════════

def docx_to_md(docx_path: str) -> str:
    """用 LibreOffice 将 DOCX 转为 HTML，再解析为 Markdown。
    保留页眉序号（一、1、（1））、标题层级、表格结构。
    """
    import tempfile, shutil
    from bs4 import BeautifulSoup, Tag

    # ── LibreOffice 转 HTML ──────────────────────────────
    tmpdir = tempfile.mkdtemp(prefix="docx_html_")
    try:
        r = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "html",
             "--outdir", tmpdir, docx_path],
            capture_output=True, timeout=120
        )
        html_files = [f for f in os.listdir(tmpdir) if f.endswith(".html")]
        if not html_files:
            raise RuntimeError("LibreOffice 未生成 HTML")
        with open(os.path.join(tmpdir, html_files[0]), encoding="utf-8", errors="ignore") as f:
            html = f.read()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    # ── HTML → Markdown ──────────────────────────────────
    soup = BeautifulSoup(html, "html.parser")

    def _cell_text(td) -> str:
        return td.get_text(separator=" ", strip=True).replace("|", "｜")

    def _tbl_to_md(tbl) -> str:
        rows = []
        for tr in tbl.find_all("tr", recursive=True):
            cells = [_cell_text(c) for c in tr.find_all(["td", "th"], recursive=False)]
            if cells:
                rows.append(cells)
        if not rows:
            return ""
        ncols = max(len(r) for r in rows)
        rows = [r + [""] * (ncols - len(r)) for r in rows]
        lines = ["| " + " | ".join(rows[0]) + " |",
                 "|" + "|".join("---" for _ in rows[0]) + "|"]
        for r in rows[1:]:
            lines.append("| " + " | ".join(r) + " |")
        return "\n".join(lines)

    blocks = []
    skip_tags = set()  # 已处理的 Tag id，避免重复

    def _process(node):
        if not isinstance(node, Tag):
            return
        if id(node) in skip_tags:
            return
        name = node.name

        if name in ("h1", "h2", "h3", "h4", "h5", "h6"):
            level = int(name[1])
            text = node.get_text(strip=True)
            if text:
                blocks.append("#" * level + " " + text)
            skip_tags.add(id(node))

        elif name == "table":
            md = _tbl_to_md(node)
            if md:
                blocks.append(md)
            skip_tags.add(id(node))

        elif name in ("ul", "ol"):
            for li in node.find_all("li", recursive=False):
                text = li.get_text(separator=" ", strip=True)
                if text:
                    blocks.append("- " + text)
                skip_tags.add(id(li))
            skip_tags.add(id(node))

        elif name == "p":
            # 跳过已被表格/列表处理过的 p
            if any(id(node) == sid for sid in skip_tags):
                return
            # 跳过 p 在 td/th/li 内（避免重复）
            parent = node.parent
            if parent and parent.name in ("td", "th", "li"):
                return
            text = node.get_text(separator=" ", strip=True)
            if text:
                blocks.append(text)

        else:
            for child in node.children:
                _process(child)

    body = soup.find("body") or soup
    for child in body.children:
        _process(child)

    return "\n\n".join(blocks)


# ═══════════════════════════════════════════════════════
#  4. XLSX → MD
# ═══════════════════════════════════════════════════════

def xlsx_to_sheets(xlsx_path: str) -> list[tuple[str, str]]:
    """返回 [(sheet_name, md_content), ...] 每个 sheet 独立。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    results = []

    for ws in wb.worksheets:
        # 展开合并单元格
        # 横向合并：只保留最左列的值，其余列留空（避免重复）
        # 纵向合并：向下填充（Markdown 不支持 rowspan）
        merge_map: dict[tuple, str] = {}
        for rng in ws.merged_cells.ranges:
            val = ws.cell(rng.min_row, rng.min_col).value
            val_str = str(val).strip() if val is not None else ""
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if c == rng.min_col:
                        merge_map[(r, c)] = val_str  # 最左列：保留值（含向下填充）
                    else:
                        merge_map[(r, c)] = ""        # 横向扩展格：留空

        rows = []
        for ri, row in enumerate(ws.iter_rows(), start=1):
            cells = []
            for ci, cell in enumerate(row, start=1):
                if (ri, ci) in merge_map:
                    cells.append(merge_map[(ri, ci)])
                else:
                    v = cell.value
                    cells.append(str(v).strip().replace("\n", " ").replace("|", "｜") if v is not None else "")
            if any(c for c in cells):
                rows.append(cells)

        if not rows:
            continue

        # 裁掉尾部全空列（ODS 转 xlsx 常带大量空列）
        ncols = max(len(r) for r in rows)
        rows = [r + [""] * (ncols - len(r)) for r in rows]
        while ncols > 0 and all(not r[ncols - 1] for r in rows):
            ncols -= 1
        if ncols == 0:
            continue
        rows = [r[:ncols] for r in rows]

        out = [f"# {ws.title}"]
        out.append("| " + " | ".join(rows[0]) + " |")
        out.append("|" + "|".join("---" for _ in rows[0]) + "|")
        for r in rows[1:]:
            out.append("| " + " | ".join(r) + " |")
        results.append((ws.title, "\n".join(out)))

    return results


# ═══════════════════════════════════════════════════════
#  5. 主流程
# ═══════════════════════════════════════════════════════

def _should_skip(rel_path: str) -> bool:
    parts = Path(rel_path).parts
    if parts[0] != '工管部':
        return True
    if parts[0] in SKIP_DOMAINS:
        return True
    if len(parts) >= 2 and parts[1] in SKIP_SUBDIRS:
        return True
    return False

def process_file(src_path: str) -> bool:
    rel  = os.path.relpath(src_path, KB_DIR)
    if _should_skip(rel):
        return False

    out_path = os.path.join(MD_DIR, os.path.splitext(rel)[0] + ".md")
    if os.path.exists(out_path):
        logger.info(f"skip (already exists): {rel}")
        return True

    logger.info(f"processing: {rel}")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    ext = os.path.splitext(src_path)[1].lower()
    try:
        if ext == ".pdf":
            content = pdf_to_md(src_path, str(Path(rel).parent))
            if not content or not content.strip():
                logger.warning(f"  empty output: {rel}")
                return False
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(content)
            logger.info(f"  ok: {out_path} ({len(content):,} chars)")
            return True

        elif ext == ".doc":
            logger.warning(f"  .doc 格式不支持（python-docx仅支持.docx），跳过: {rel}")
            return False

        elif ext == ".docx":
            content = docx_to_md(src_path)
            if not content or not content.strip():
                logger.warning(f"  empty output: {rel}")
                return False
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(content)
            logger.info(f"  ok: {out_path} ({len(content):,} chars)")
            return True

        elif ext in (".xlsx", ".xls", ".ods"):
            actual_path = src_path
            if ext == ".ods":
                logger.info(f"  .ods → .xlsx 转换中...")
                _ods_out = os.path.join(CACHE_DIR, "_ods_converted")
                os.makedirs(_ods_out, exist_ok=True)
                subprocess.run(
                    ["libreoffice", "--headless", "--convert-to", "xlsx",
                     "--outdir", _ods_out, src_path],
                    capture_output=True, timeout=300,
                )
                _base = os.path.splitext(os.path.basename(src_path))[0]
                actual_path = os.path.join(_ods_out, _base + ".xlsx")
                if not os.path.exists(actual_path):
                    logger.error(f"  ODS 转换失败: {rel}")
                    return False
            sheets = xlsx_to_sheets(actual_path)
            if not sheets:
                logger.warning(f"  empty output: {rel}")
                return False
            base_no_ext = os.path.splitext(out_path)[0]
            if len(sheets) == 1:
                # 单 sheet 直接用原文件名
                with open(out_path, "w", encoding="utf-8") as f:
                    f.write(sheets[0][1])
                logger.info(f"  ok: {out_path}")
            else:
                # 多 sheet：每个 sheet 一个文件，放在同名子目录下
                os.makedirs(base_no_ext, exist_ok=True)
                for sheet_name, md in sheets:
                    safe_name = re.sub(r'[\\/:*?"<>|]', '_', sheet_name)
                    sheet_path = os.path.join(base_no_ext, f"{safe_name}.md")
                    with open(sheet_path, "w", encoding="utf-8") as f:
                        f.write(md)
                    logger.info(f"  ok: {sheet_path} ({len(md):,} chars)")
            return True

        elif ext == ".md":
            import shutil
            shutil.copy2(src_path, out_path)
            logger.info(f"  ok (copied): {out_path}")
            return True

        else:
            logger.warning(f"  unsupported: {ext}")
            return False

    except Exception as e:
        logger.error(f"  failed: {e}")
        return False


def main():
    exts = ("*.pdf", "*.PDF", "*.docx", "*.DOCX", "*.doc", "*.DOC",
            "*.xlsx", "*.XLSX", "*.xls", "*.XLS", "*.ods", "*.ODS",
            "*.md", "*.MD")
    all_files = []
    for ext in exts:
        all_files.extend(glob.glob(os.path.join(KB_DIR, "**", ext), recursive=True))
    all_files = sorted(
        f for f in all_files
        if not _should_skip(str(Path(f).relative_to(KB_DIR)))
    )

    logger.info(f"total files: {len(all_files)}")
    ok = fail = 0
    for f in all_files:
        if process_file(f):
            ok += 1
        else:
            fail += 1
    logger.info(f"done: {ok} ok, {fail} failed/skipped")


if __name__ == "__main__":
    main()
